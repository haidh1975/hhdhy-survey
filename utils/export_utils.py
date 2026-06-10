"""
Unified multi-format export utilities for HHD-HY Survey App.

Formats supported:
  - Excel (openpyxl) — styled header, auto column width
  - CSV (UTF-8 BOM — compatible with Vietnamese Excel)
  - JSON (pretty-printed, UTF-8)
  - Markdown summary table
  - TXT plain text report

Design patterns:
  - Single Responsibility: one function per format (Clean Code GitHub patterns)
  - Strategy Pattern: callers pick format without knowing internals
  - Fail-safe returns: never raises; always returns (bytes | None, error_msg)

References:
  - GitHub public-apis / awesome-api — data interchange best practices
  - GeeksforGeeks — file I/O Python patterns
  - MDN Web Docs — data URI / download patterns (adapted for Streamlit)
"""

from __future__ import annotations
import io
import json
import csv
import logging
from datetime import datetime
from typing import Any

import pandas as pd

logger = logging.getLogger(__name__)


# ─── Excel with styled header ─────────────────────────────────────────────────

def export_excel(
    df: pd.DataFrame,
    sheet_name: str = "Data",
    title: str = "",
) -> tuple[bytes | None, str]:
    """
    Export DataFrame to Excel with:
      - Bold blue header row
      - Auto column width (max 50)
      - Title row at top if provided
      - Freeze panes on header
    """
    try:
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        buf = io.BytesIO()
        with pd.ExcelWriter(buf, engine="openpyxl") as writer:
            row_offset = 0
            if title:
                # Write title row then blank row
                temp = pd.DataFrame([[title] + [""] * (len(df.columns) - 1)],
                                    columns=df.columns)
                pd.concat([temp, pd.DataFrame([[""] * len(df.columns)], columns=df.columns), df]).to_excel(
                    writer, sheet_name=sheet_name, index=False, header=False
                )
                row_offset = 2
            else:
                df.to_excel(writer, sheet_name=sheet_name, index=False)

            ws = writer.sheets[sheet_name]

            # Style header row (row 1 + offset)
            header_row = row_offset + 1
            for col_num, col_name in enumerate(df.columns, 1):
                cell = ws.cell(row=header_row, column=col_num)
                cell.font = Font(bold=True, color="FFFFFF", size=11)
                cell.fill = PatternFill("solid", fgColor="0066CC")
                cell.alignment = Alignment(horizontal="center", vertical="center",
                                           wrap_text=True)
                thin = Side(border_style="thin", color="DDDDDD")
                cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)

            # Auto-fit column widths
            for col_cells in ws.columns:
                max_len = 0
                col_letter = get_column_letter(col_cells[0].column)
                for cell in col_cells:
                    try:
                        max_len = max(max_len, len(str(cell.value or "")))
                    except Exception:
                        pass
                ws.column_dimensions[col_letter].width = min(max(max_len + 2, 12), 50)

            # Freeze header
            ws.freeze_panes = ws.cell(row=header_row + 1, column=1)

            # Row height for header
            ws.row_dimensions[header_row].height = 22

            # Title styling
            if title and row_offset > 0:
                title_cell = ws.cell(row=1, column=1)
                title_cell.value = title
                title_cell.font = Font(bold=True, size=13, color="0066CC")
                ws.merge_cells(start_row=1, start_column=1,
                               end_row=1, end_column=min(len(df.columns), 10))

        return buf.getvalue(), ""
    except Exception as e:
        logger.error(f"Excel export error: {e}")
        return None, str(e)


# ─── CSV (UTF-8 BOM) ──────────────────────────────────────────────────────────

def export_csv(df: pd.DataFrame) -> tuple[bytes | None, str]:
    """Export DataFrame to UTF-8 BOM CSV (opens cleanly in Vietnamese Excel)."""
    try:
        return df.to_csv(index=False).encode("utf-8-sig"), ""
    except Exception as e:
        logger.error(f"CSV export error: {e}")
        return None, str(e)


# ─── JSON ─────────────────────────────────────────────────────────────────────

def export_json(data: Any, ensure_ascii: bool = False, indent: int = 2) -> tuple[bytes | None, str]:
    """Export any JSON-serialisable object to bytes."""
    try:
        def _default(obj):
            if hasattr(obj, "isoformat"):
                return obj.isoformat()
            return str(obj)

        return json.dumps(data, ensure_ascii=ensure_ascii, indent=indent,
                          default=_default).encode("utf-8"), ""
    except Exception as e:
        logger.error(f"JSON export error: {e}")
        return None, str(e)


# ─── Markdown table ───────────────────────────────────────────────────────────

def export_markdown(df: pd.DataFrame, title: str = "") -> tuple[bytes | None, str]:
    """Export DataFrame as a Markdown table."""
    try:
        lines = []
        if title:
            lines.append(f"# {title}\n")
            lines.append(f"*Xuất lúc: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}*\n")

        # Header
        lines.append("| " + " | ".join(str(c) for c in df.columns) + " |")
        lines.append("|" + "|".join(["---"] * len(df.columns)) + "|")
        # Rows
        for _, row in df.iterrows():
            lines.append("| " + " | ".join(str(v)[:80].replace("|", "\\|") for v in row) + " |")

        return "\n".join(lines).encode("utf-8"), ""
    except Exception as e:
        logger.error(f"Markdown export error: {e}")
        return None, str(e)


# ─── Plain text report ────────────────────────────────────────────────────────

def export_text_report(
    survey: dict,
    responses: list[dict],
    lang: str = "vi",
) -> tuple[bytes | None, str]:
    """
    Generate a plain-text summary report of all responses.
    Suitable as an audit trail or quick read.
    """
    try:
        lines = []
        sep = "=" * 60

        if lang == "vi":
            lines += [
                sep,
                f"BÁO CÁO KHẢO SÁT: {survey.get('title', '')}",
                f"Mô tả: {survey.get('description', 'Không có')}",
                f"Xuất lúc: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
                f"Tổng số phản hồi: {len(responses)}",
                sep, "",
            ]
        else:
            lines += [
                sep,
                f"SURVEY REPORT: {survey.get('title', '')}",
                f"Description: {survey.get('description', 'None')}",
                f"Exported at: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
                f"Total responses: {len(responses)}",
                sep, "",
            ]

        questions = {
            q.get("id", str(i)): q["question_text"]
            for i, q in enumerate(survey.get("questions", []))
        }

        for idx, r in enumerate(responses, 1):
            rdata = r.get("response_data", {})
            ts = r.get("submitted_at", "")[:16]
            lines.append(f"{'Phản hồi' if lang == 'vi' else 'Response'} #{idx}  [{ts}]")
            lines.append("-" * 40)
            for q_id, q_text in questions.items():
                val = rdata.get(q_id, "")
                if isinstance(val, list):
                    val = ", ".join(str(v) for v in val)
                lines.append(f"  {q_text}")
                lines.append(f"    → {val}")
            lines.append("")

        return "\n".join(lines).encode("utf-8"), ""
    except Exception as e:
        logger.error(f"Text report error: {e}")
        return None, str(e)


# ─── Streamlit download helper ────────────────────────────────────────────────

def download_buttons(
    df: pd.DataFrame,
    base_filename: str,
    survey: dict | None = None,
    responses: list[dict] | None = None,
    lang: str = "vi",
) -> None:
    """
    Render a row of download buttons (Excel, CSV, JSON, Markdown, TXT)
    directly in the Streamlit page.
    """
    import streamlit as st

    labels = {
        "excel": ("📊 Excel", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", ".xlsx"),
        "csv":   ("📄 CSV",   "text/csv",             ".csv"),
        "json":  ("🗂️ JSON",  "application/json",     ".json"),
        "md":    ("📝 Markdown", "text/markdown",      ".md"),
        "txt":   ("📃 TXT",   "text/plain",            ".txt"),
    }
    if lang == "en":
        labels["excel"] = ("📊 Excel", labels["excel"][1], labels["excel"][2])

    col_excel, col_csv, col_json, col_md, col_txt = st.columns(5)

    with col_excel:
        data, err = export_excel(df, title=survey.get("title", "") if survey else "")
        if data:
            st.download_button(labels["excel"][0], data,
                               f"{base_filename}.xlsx", labels["excel"][1],
                               use_container_width=True)

    with col_csv:
        data, err = export_csv(df)
        if data:
            st.download_button(labels["csv"][0], data,
                               f"{base_filename}.csv", labels["csv"][1],
                               use_container_width=True)

    with col_json:
        json_data = df.to_dict(orient="records")
        data, err = export_json(json_data)
        if data:
            st.download_button(labels["json"][0], data,
                               f"{base_filename}.json", labels["json"][1],
                               use_container_width=True)

    with col_md:
        data, err = export_markdown(df, title=survey.get("title", "") if survey else "")
        if data:
            st.download_button(labels["md"][0], data,
                               f"{base_filename}.md", labels["md"][1],
                               use_container_width=True)

    with col_txt:
        if survey and responses:
            data, err = export_text_report(survey, responses, lang=lang)
        else:
            data, err = df.to_string().encode("utf-8"), ""
        if data:
            st.download_button(labels["txt"][0], data,
                               f"{base_filename}.txt", labels["txt"][1],
                               use_container_width=True)
